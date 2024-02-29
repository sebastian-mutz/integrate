"""
author: daniel boateng (dannboateng@gmail.com)
This script basically controls the measurement from BME280 sensor, store in an excel sheet (with defined columns for date, time, temperature,
humidity and pressure. Not that other parmaters like dewpoint temperature can be calculated from these parameters) and also displays the date
and time on attached oled screen for each measurement.
"""

# The first part of the script is to import all the modules we installed at the previous section required for the devices. In python, 
# "import" is used to call a package in current script. 

#importing modules 
import smbus2
import bme280
import time
import datetime
from datetime import date
from openpyxl import load_workbook
import board
import busio
import digitalio
import adafruit_ssd1306
from PIL import Image, ImageDraw, ImageFont


# We will set the bme280 object in order use its functions for measurement

# define I2C port (normally, I2C has 2 ports (i.e., I2C0, I2C1)). So if you use "i2cdetect -y 1" to check  
# address of your device, then the devices are on port 1. Likewise "i2cdetect -y 0" if they are connected to port 0
port = 1

# define the address of sensor (as explained from the previous section using "i2cdetect")
address = 0x76

# define SMbus instance for controlling the board
bus = smbus2.SMBus(port)

# Now we will calibrate the sensor using the load_calibrations function, which is attribute of bme280 module
# by using the bus and address as input parameters
calibration_params = bme280.load_calibration_params(bus, address)

# define data class to calculate the readings from the sensor (with the calibrated params)
data = bme280.sample(bus, address, calibration_params)

# Now each variable can be obtained from the "data class" by calling variable method. But we will take the 
# firt readings and ignore it since it might not be correct. Then after 1 seconds (this can be increased)
# the correct readings can be recored on a specific interval.
temperature = data.temperature
pressure = data.pressure
humidity = data.humidity

time.sleep(1)   # time interval to wait after the first reading

# setting OLED screen 

# define reset pins by using digit input and output (the board can be any number eg. D5, D6..)
oled_reset = digitalio.DigitalInOut(board.D4)

# define the right size of your display. Please change it depending on your display size (eg. 128x32)
WIDTH = 128
HEIGHT = 64
BORDER = 5

# define I2C
i2c = board.I2C()

# define the oled display class (using the address and defined paramters)
oled = adafruit_ssd1306.SSD1306_I2C(WIDTH, HEIGHT, i2c, addr=0x3c, reset=oled_reset)


# Now lets clear the display first
oled.fill(0)
oled.show()

# customizing the screen specifications (we are using default here)
width = oled.width
height = oled.height

# now create blank image to draw your text on it. Not that "1" is for 1-bit color
image = Image.new("1", (width, height))

#define object for drawing on the image 
draw =ImageDraw.Draw(image)

# define a black backgroud to draw the image on (note you can use 225 for fill and outline
# for white background)
draw.rectangle((0,0, width, height), outline=0, fill=0)

# define the positon of the text
padding=-2
top=padding
bottom=height-padding
x=0

#define font type (we used default here, you can google it for other styles)
font = ImageFont.load_default()

#saving data

# Now we will load the excel file created on Desktop into memory by using the load_workbook from openpyxl

wb = load_workbook("/home/pi/Desktop/Reading_from_bme280.xlsx")
sheet = wb["Sheet1"]

# Now we will write a loop to continue the measurement until we stop by pressing Ctrl + C

print("Please to end the readings from the sensor, kindly press Ctrl + C") # print it for new users to know

# So this loop will run until you interupt it (But you can also define it to stop after certain time)
try:

    while True:
        draw.rectangle((0,0, width, height), outline=0, fill=0)

        # taking readings from the data class and rounding it once decimal place
        temperature = round(data.temperature,1)
        pressure = round(data.pressure,1)
        humidity = round(data.humidity,1)

        # check date for measuring 
        today = date.today()

        # time of measurement
        now = datetime.datetime.now().time()

        print("Adding data to the spreadsheet")
        print(today)
        print(now)
        print("{}°C {}hPa {}%".format(temperature, pressure, humidity))
        
        # define the data to be stored in the excel fine on each row (so define the excel as similar !!)
        row = (today, now, temperature, pressure, humidity)
        sheet.append(row)
        wb.save("/home/pi/Desktop/Reading_from_bme280.xlsx")  # saving after each measurement

   
        # drawing text on the screen with their positions
        draw.text((x,top+0), str(today), font=font, fill=255)
        draw.text((x, top+20), str(now), font=font, fill=255)
        draw.text((x, top+40), "{}°C {}hPa {}%".format(temperature, pressure, humidity), font=font, fill=255)
        oled.image(image)
        oled.show()

        # define waiting time after each measurement (eg. 60 seconds)
        time.sleep(60)

# now make sure it saves everyting after stoping the reading !        
finally:
    wb.save("/home/pi/Desktop/Reading_from_bme280.xlsx")
    print("Goodbye!")


   
